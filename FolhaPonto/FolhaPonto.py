from os.path import exists
from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta
from tkinter import filedialog
from FolhaPontoMODELS import fullDay
from openpyxl.styles import Font
from openpyxl.worksheet.dimensions import RowDimension
import openpyxl
import random


csv_file = ''
def verify_if_csv_file_exsists():
    if exists(csv_file):
        tabela_horarios = open(csv_file, 'a')
        return tabela_horarios
    else:
        create_csv_file()


def create_csv_file():
    tabela_horarios = open(csv_file, 'w')
    return tabela_horarios


def format_into_hour_minute(valor_in_minutes):
    if valor_in_minutes != '':
        hour = valor_in_minutes//60
        minutes = int((valor_in_minutes % 60))
        hour = str(hour).zfill(2)
        minutes = str(minutes).zfill(2)
        return f'{hour}:{minutes}'
    return valor_in_minutes

def ask_for_csv_file_name_with_path():
    global csv_file
    if len(csv_file) == 0:
        csv_file = filedialog.asksaveasfilename()
    return csv_file


def convert_str_to_date(date_str):
    date = datetime.strptime(date_str, '%Y-%m-%d').date()
    return date


def calculate_end_date(initial_date, date_of_end):
    if date_of_end == '':
        return initial_date + relativedelta(months=1) - relativedelta(days=1)
    return convert_str_to_date(date_of_end)


def assemble_days_list(day, final_date):
    list_of_dates = []
    while day <= final_date:
        list_of_dates.append(day)
        day += timedelta(days=1)
    return list_of_dates


def convert_strdate_to_intminutes(date_str):
    return int(date_str.split(':')[1]) + (int(date_str.split(':')[0])*60)


def randomizing_schedules(dict_config_in_minutes, date):
    date["start_hour"] = round(random.randint((dict_config_in_minutes["begin_hour"] - dict_config_in_minutes["tolerance"]),
                                              (dict_config_in_minutes["begin_hour"] + dict_config_in_minutes["tolerance"])), 2)
    date["out_to_break"] = round(random.randint((dict_config_in_minutes["begin_lunch"] - dict_config_in_minutes["tolerance"]),
                                                (dict_config_in_minutes["begin_lunch"] + dict_config_in_minutes["tolerance"])), 2)
    date["in_from_break"] = round(random.randint((dict_config_in_minutes["end_lunch"] - dict_config_in_minutes["tolerance"]),
                                                 (dict_config_in_minutes["end_lunch"] + dict_config_in_minutes["tolerance"])), 2)
    date["stop_hour"] = round(random.randint((dict_config_in_minutes["end_hour"] - dict_config_in_minutes["tolerance"]),
                                             (dict_config_in_minutes["end_hour"] + dict_config_in_minutes["tolerance"])), 2)
    return date

def convert_standard_schedules_in_minutes(dict_config):
    dict_config["tolerance"] = int(dict_config["tolerance"])
    dict_config["begin_hour"] = convert_strdate_to_intminutes(dict_config["begin_hour"])
    dict_config["begin_lunch"] = convert_strdate_to_intminutes(dict_config["begin_lunch"])
    dict_config["end_lunch"] = convert_strdate_to_intminutes(dict_config["end_lunch"])
    dict_config["end_hour"] = convert_strdate_to_intminutes(dict_config["end_hour"])
    return dict_config


def validate_acptable_schedules(dict_config_in_minutes, date):
    if (dict_config_in_minutes["standard_work_hours"] + dict_config_in_minutes["tolerance"]) > (date["stop_hour"] - date["in_from_break"]) + \
            (date["out_to_break"] - date["start_hour"]) > (dict_config_in_minutes["standard_work_hours"] - dict_config_in_minutes["tolerance"]) and \
            ((date["stop_hour"] - date["in_from_break"]) + (date["out_to_break"] - date["start_hour"])) % 60 != 0:
        return True
    return False


def calculate_day_working_schedules(dict_config_in_minutes, date):
    print(dict_config_in_minutes["has_saturday"])
    if date["date"].weekday() > dict_config_in_minutes["has_saturday"]:
        date["start_hour"] = ''
        date["out_to_break"] = ''
        date["in_from_break"] = ''
        date["stop_hour"] = ''
        return date
    date = randomizing_schedules(dict_config_in_minutes, date)
    while not validate_acptable_schedules(dict_config_in_minutes, date):
            date = randomizing_schedules(dict_config_in_minutes, date)
    return date

def assemble_period_schedules(dict_config, list_of_days):
    date = {}
    list_of_dates = []
    dict_config_in_minutes = convert_standard_schedules_in_minutes(dict_config)
    dict_config_in_minutes["standard_work_hours"] = (dict_config_in_minutes["begin_lunch"]-dict_config_in_minutes["begin_hour"]) + \
                          (dict_config_in_minutes["end_hour"]- dict_config_in_minutes["end_lunch"])

    for day in list_of_days:

        date["date"] = day
        date = calculate_day_working_schedules(dict_config_in_minutes, date)
        date["date"] = day.strftime('%d/%m/%Y')
        date["start_hour"] = format_into_hour_minute(date["start_hour"])
        date["out_to_break"] = format_into_hour_minute(date["out_to_break"])
        date["in_from_break"] = format_into_hour_minute(date["in_from_break"])
        date["stop_hour"] = format_into_hour_minute(date["stop_hour"])
        print(date)
        date_obj = fullDay(date)
        list_of_dates.append(date_obj)
    return list_of_dates


def write_to_csv(linha_horarios):

    tabela_horarios = verify_if_csv_file_exsists()
    breakpoint()
    tabela_horarios.write(linha_horarios)
    tabela_horarios.close()

def monta_tabela_horarios(day, data_fim):

    while day <= data_fim:
        day_to_write = day.strftime('%d/%m/%Y')
        line_to_write = f'{day_to_write};'
        print(line_to_write)
        if 5 > day.weekday():

            linha_horarios = calculate_day_working_schedules()
            line_to_write = line_to_write + linha_horarios
            write_to_csv(line_to_write)

            day = day + timedelta(days=1)

        else:
            linha_horarios = '\n'
            line_to_write = line_to_write + linha_horarios
            write_to_csv(line_to_write)
            day = day + timedelta(days=1)
    

'''def header_assembler(data_inicio, data_fim):
    wb = openpyxl.Workbook()
    xlsx_file = 'FolhaPonto.xlsx'

    nome, funcao, dict_company = get_user_config()
    print(dict_company.keys())
    mes_ano = (data_inicio + relativedelta(months=1)).strftime('%b-%y')
    ws = wb.create_sheet(f'{mes_ano}')

    print(ws.row_dimensions)
    ws.calculate_dimension()
    data_inicio = data_inicio.strftime('%d/%m/%Y')
    data_fim = data_fim.strftime('%d/%m/%Y')

    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 13

    line_1_font = Font(name='Calibri', size = 18, bold = True, color = 'FF000000')
    ws['A1'].font = line_1_font
    ws['A1'] = 'FOLHA PONTO'

    ws['D1'].font = line_1_font
    ws['D1'] = 'CNPJ :'

    ws['E1'].font = line_1_font
    ws['E1'] = f"{dict_company['cnpj']}"


    line_2_font = Font(name='Calibri', size=16, bold=True, color='FF000000')
    ws['A2'].font = line_2_font
    ws['A2'] = f"{dict_company['nome_fantasia']}"

    line_3_font = Font(name='Calibri', size=14, bold=True, color='FF000000')
    ws['A3'].font = line_3_font
    ws['A3'] = f"{'Nome'} : {dict_company['nome']}"

    ws['E3'].font = line_3_font
    ws['E3'] = f"{'Funç'} : {dict_company['funcao']}"

    line_4_font = Font(name='Calibri', size=14, bold=True, color='FF000000')
    ws['A4'].font = line_4_font
    ws['A4'] = f"{mes_ano}"

    ws['D4'].font = line_4_font
    ws['D4'] = f"{data_inicio} à {data_fim}"

    line_5_font = Font(name='Calibri', size=12, bold=True, color='FF000000')
    ws['A5'].font = line_5_font
    ws['A5'] = 'Data'
    ws['B5'].font = line_5_font
    ws['B5'] = 'Horário de\nEntrada'
    ws['C5'].font = line_5_font
    ws['C5'] = 'Entrada no\nIntervalo'
    ws['D5'].font = line_5_font
    ws['D5'] = 'Saída do\nIntervalo'
    ws['E5'].font = line_5_font
    ws['E5'] = 'Hoarario de\nSaída'

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 40

    wb.save(xlsx_file)
    return cabecalho


def get_user_config():
    config = open('UserConfig.txt', 'r').readlines()
    nome = config[0].replace('nome:', '').replace('\n', '').strip()
    funcao = config[1].replace('função:', '').replace('\n', '').strip()
    dict_company = {(config[0].split(':')[0]):(config[0].split(':')[1].replace('\n', '').strip()),
                    (config[1].split(':')[0]):(config[1].split(':')[1].replace('\n', '').strip()),
                    (config[2].split(':')[0]):(config[2].split(':')[1].replace('\n', '').strip()),
                    (config[3].split(':')[0]):(config[3].split(':')[1].replace('\n', '').strip())}
    return nome, funcao, dict_company'''

def main():

    
    day = convert_str_to_date()
    data_fim = calculate_end_date(day)
    #cabecalho = header_assembler(fullDay,data_fim)
    ask_for_csv_file_name_with_path()
    #write_to_csv(cabecalho)
    monta_tabela_horarios(day,data_fim)
    #footer = footer_assembler()
    #write_to_csv(footer)
    exit()

#main()
